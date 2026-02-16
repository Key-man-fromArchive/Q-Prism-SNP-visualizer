"""Parser for Bio-Rad CFX Opus Quantification Amplification Results .xlsx files.

File structure (after fixing broken packaging):
- Sheets: FAM, HEX, ROX, Run Information
- Each dye sheet is WIDE format:
  - Column A: empty
  - Column B: Cycle number
  - Columns C-CT: Well IDs as headers (A1, A2, ..., H12)
  - Row 1: Header row
  - Row 2+: One row per cycle (this sample has only 1 cycle)
- Well IDs are native A1-H12 format
"""

import os

import openpyxl

from app.models import UnifiedData, WellCycleData
from app.parsers.xlsx_fixer import fix_cfx_xlsx, needs_fixing


def parse_cfx_opus(file_path: str) -> UnifiedData:
    # Fix broken xlsx if needed
    fixed_path = None
    if needs_fixing(file_path):
        fixed_path = fix_cfx_xlsx(file_path)
        work_path = fixed_path
    else:
        work_path = file_path

    try:
        wb = openpyxl.load_workbook(work_path, data_only=True)
        return _parse_workbook(wb)
    finally:
        wb.close()
        if fixed_path and os.path.exists(fixed_path):
            os.remove(fixed_path)


def _parse_workbook(wb: openpyxl.Workbook) -> UnifiedData:
    sheet_names = wb.sheetnames

    # Detect allele2 dye
    allele2_dye = "HEX" if "HEX" in sheet_names else "VIC"
    has_rox = "ROX" in sheet_names

    # Parse each dye sheet into {(well, cycle): value}
    fam_data = _parse_dye_sheet(wb["FAM"])
    allele2_data = _parse_dye_sheet(wb[allele2_dye])
    rox_data = _parse_dye_sheet(wb["ROX"]) if has_rox else {}

    # Merge into unified data
    all_keys = set(fam_data.keys()) | set(allele2_data.keys())

    wells_set: set[str] = set()
    cycles_set: set[int] = set()
    data: list[WellCycleData] = []

    for well, cycle in sorted(all_keys):
        fam_val = fam_data.get((well, cycle), 0.0)
        allele2_val = allele2_data.get((well, cycle), 0.0)
        rox_val = rox_data.get((well, cycle)) if has_rox else None

        data.append(
            WellCycleData(
                well=well,
                cycle=cycle,
                fam=fam_val,
                allele2=allele2_val,
                rox=rox_val,
            )
        )
        wells_set.add(well)
        cycles_set.add(cycle)

    return UnifiedData(
        instrument="CFX Opus",
        allele2_dye=allele2_dye,
        wells=sorted(wells_set, key=_well_sort_key),
        cycles=sorted(cycles_set),
        data=data,
        has_rox=has_rox,
    )


def _parse_dye_sheet(ws) -> dict[tuple[str, int], float]:
    """Parse a wide-format dye sheet into {(well_id, cycle): rfu_value}."""
    result = {}

    # Row 1 is header: [None, 'Cycle', 'A1', 'A2', ..., 'H12']
    headers = [cell.value for cell in ws[1]]

    # Well IDs start at column index 2 (column C)
    well_ids = []
    for i in range(2, len(headers)):
        h = headers[i]
        if h is not None:
            well_ids.append((i, str(h)))

    # Data rows start at row 2
    for row_idx in range(2, ws.max_row + 1):
        cycle_val = ws.cell(row=row_idx, column=2).value
        if cycle_val is None:
            continue
        cycle = int(cycle_val)

        for col_idx, well_id in well_ids:
            val = ws.cell(row=row_idx, column=col_idx + 1).value
            if val is not None and isinstance(val, (int, float)):
                result[(well_id, cycle)] = float(val)

    return result


def _open_cfx(file_path: str):
    """Open a CFX Opus file, fixing packaging if needed. Returns (wb, fixed_path)."""
    fixed_path = None
    if needs_fixing(file_path):
        fixed_path = fix_cfx_xlsx(file_path)
        work_path = fixed_path
    else:
        work_path = file_path
    wb = openpyxl.load_workbook(work_path, data_only=True)
    return wb, fixed_path


def parse_cfx_endpoint(file_path: str) -> UnifiedData:
    """Parse CFX Opus End Point Results .xlsx.

    Structure: FAM/HEX/ROX sheets, each in long format:
    [None, Well, Fluor, Target, Content, Sample, End RFU, Call, Sample Type, CallType, Is Control]
    96 rows per sheet (one per well).
    """
    wb, fixed_path = _open_cfx(file_path)
    try:
        sheet_names = wb.sheetnames
        allele2_dye = "HEX" if "HEX" in sheet_names else "VIC"
        has_rox = "ROX" in sheet_names

        fam_data = _parse_endpoint_sheet(wb["FAM"])
        allele2_data = _parse_endpoint_sheet(wb[allele2_dye])
        rox_data = _parse_endpoint_sheet(wb["ROX"]) if has_rox else {}

        all_wells = set(fam_data.keys()) | set(allele2_data.keys())
        data: list[WellCycleData] = []
        wells_set: set[str] = set()

        for well in sorted(all_wells):
            fam_val = fam_data.get(well, 0.0)
            allele2_val = allele2_data.get(well, 0.0)
            rox_val = rox_data.get(well) if has_rox else None

            data.append(WellCycleData(
                well=well, cycle=1,
                fam=fam_val, allele2=allele2_val, rox=rox_val,
            ))
            wells_set.add(well)

        return UnifiedData(
            instrument="CFX Opus",
            allele2_dye=allele2_dye,
            wells=sorted(wells_set, key=_well_sort_key),
            cycles=[1],
            data=data,
            has_rox=has_rox,
        )
    finally:
        wb.close()
        if fixed_path and os.path.exists(fixed_path):
            os.remove(fixed_path)


def _parse_endpoint_sheet(ws) -> dict[str, float]:
    """Parse End Point Results sheet: {well_id: end_rfu}."""
    result = {}
    headers = [cell.value for cell in ws[1]]
    headers_str = [str(h).upper() if h else "" for h in headers]

    well_col = None
    rfu_col = None
    for i, h in enumerate(headers_str):
        if h == "WELL":
            well_col = i
        if h == "END RFU":
            rfu_col = i

    if well_col is None or rfu_col is None:
        return result

    for row_idx in range(2, ws.max_row + 1):
        well = ws.cell(row=row_idx, column=well_col + 1).value
        rfu = ws.cell(row=row_idx, column=rfu_col + 1).value

        if well and rfu is not None and isinstance(rfu, (int, float)):
            # Normalize well format: A01 -> A1
            well_str = str(well).strip()
            if len(well_str) == 3 and well_str[1] == "0":
                well_str = well_str[0] + well_str[2]
            result[well_str] = float(rfu)

    return result


def parse_cfx_allelic(file_path: str) -> UnifiedData:
    """Parse CFX Opus Allelic Discrimination Results .xlsx.

    Structure: ADSheet with [None, Well, Sample, Call, Type, RFU1, RFU2]
    RFU1 = FAM, RFU2 = HEX (based on value comparison with known data).
    No ROX data available.
    """
    wb, fixed_path = _open_cfx(file_path)
    try:
        ws = wb["ADSheet"]
        headers = [cell.value for cell in ws[1]]
        headers_str = [str(h).upper() if h else "" for h in headers]

        well_col = None
        rfu1_col = None
        rfu2_col = None
        call_col = None
        sample_col = None

        for i, h in enumerate(headers_str):
            if h == "WELL":
                well_col = i
            elif h == "RFU1":
                rfu1_col = i
            elif h == "RFU2":
                rfu2_col = i
            elif h == "CALL":
                call_col = i
            elif h == "SAMPLE":
                sample_col = i

        if well_col is None or rfu1_col is None or rfu2_col is None:
            raise ValueError("ADSheet missing required columns (Well, RFU1, RFU2)")

        data: list[WellCycleData] = []
        wells_set: set[str] = set()
        sample_names: dict[str, str] = {}

        for row_idx in range(2, ws.max_row + 1):
            well = ws.cell(row=row_idx, column=well_col + 1).value
            rfu1 = ws.cell(row=row_idx, column=rfu1_col + 1).value
            rfu2 = ws.cell(row=row_idx, column=rfu2_col + 1).value

            if not well or not isinstance(rfu1, (int, float)) or not isinstance(rfu2, (int, float)):
                continue

            well_str = str(well).strip()
            if len(well_str) == 3 and well_str[1] == "0":
                well_str = well_str[0] + well_str[2]

            data.append(WellCycleData(
                well=well_str, cycle=1,
                fam=float(rfu1), allele2=float(rfu2), rox=None,
            ))
            wells_set.add(well_str)

            if sample_col is not None:
                sample = ws.cell(row=row_idx, column=sample_col + 1).value
                if sample:
                    sample_names[well_str] = str(sample)

        return UnifiedData(
            instrument="CFX Opus",
            allele2_dye="HEX",
            wells=sorted(wells_set, key=_well_sort_key),
            cycles=[1],
            data=data,
            has_rox=False,
            sample_names=sample_names if sample_names else None,
        )
    finally:
        wb.close()
        if fixed_path and os.path.exists(fixed_path):
            os.remove(fixed_path)


def _well_sort_key(well: str) -> tuple[int, int]:
    row = ord(well[0]) - ord("A")
    col = int(well[1:])
    return (row, col)
