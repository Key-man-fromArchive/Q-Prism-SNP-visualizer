"""Auto-detect instrument format and file type, with clear error messages."""

import os
import zipfile

import xlrd

from app.models import UnifiedData

# QuantStudio valid file types (by header columns)
QS_VALID_HEADERS = {
    "multicomponent": {"WELL", "CYCLE", "FAM"},  # must have FAM + (VIC or HEX)
    "amplification": {"WELL", "CYCLE", "TARGET NAME", "RN", "DELTA RN"},
}

QS_KNOWN_FILES = {
    "multicomponent data": "Multicomponent Data",
    "amplification data": "Amplification Data",
    "raw data": "Raw Data",
    "results": "Results",
    "sample setup": "Sample Setup",
}

# CFX Opus valid sheet patterns
CFX_VALID_PATTERNS = {
    "amplification": "Quantification Amplification Results",
    "endpoint": "End Point Results",
    "allelic": "Allelic Discrimination Results",
}

CFX_KNOWN_FILES = {
    "quantification amplification": "Quantification Amplification Results",
    "end point": "End Point Results",
    "allelic discrimination": "Allelic Discrimination Results",
    "quantification cq": "Quantification Cq Results",
    "quantification plate view": "Quantification Plate View Results",
    "quantification summary": "Quantification Summary",
    "melt curve": "Melt Curve Plate View Results",
    "gene expression": "Gene Expression Results",
    "anova": "ANOVA Results",
    "standard curve": "Standard Curve Results",
}


def detect_and_parse(file_path: str, original_filename: str = "") -> UnifiedData:
    ext = os.path.splitext(file_path)[1].lower()
    filename = (original_filename or os.path.basename(file_path)).lower()

    if ext == ".zip":
        return _handle_zip(file_path, filename)
    elif ext == ".eds":
        return _handle_eds(file_path)
    elif ext == ".pcrd":
        raise ValueError(
            "Bio-Rad .pcrd files are password-encrypted and cannot be read directly.\n\n"
            "Please export from CFX Maestro instead:\n"
            "  File > Export > 'Quantification Amplification Results' (.xlsx)"
        )
    elif ext == ".xls":
        return _handle_quantstudio(file_path, filename)
    elif ext == ".xlsx":
        return _handle_cfx_opus(file_path, filename)
    else:
        raise ValueError(
            f"Unsupported file extension: {ext}.\n"
            "Upload .eds (QuantStudio raw), .xls (QuantStudio export), "
            ".xlsx (CFX Opus export), or .zip (CFX XML export) files."
        )


def _handle_eds(file_path: str) -> UnifiedData:
    """Handle QuantStudio .eds raw instrument files."""
    from app.parsers.eds_raw import parse_eds

    if not zipfile.is_zipfile(file_path):
        raise ValueError(
            "This .eds file appears to be corrupted (not a valid ZIP archive).\n"
            "Try re-exporting from QuantStudio."
        )
    return parse_eds(file_path)


def _handle_zip(file_path: str, filename: str) -> UnifiedData:
    """Handle ZIP files — CFX Opus XML export archives."""
    from app.parsers.cfx_xml_parser import parse_cfx_xml_zip

    if not zipfile.is_zipfile(file_path):
        raise ValueError(
            "This .zip file appears to be corrupted (not a valid ZIP archive)."
        )

    # Quick check: does this ZIP contain any CFX XML patterns?
    with zipfile.ZipFile(file_path, "r") as zf:
        names = zf.namelist()
        has_cfx_xml = any(
            n.endswith("_ADSheet.xml") or
            n.endswith("_FAM.xml") or
            n.endswith("_HEX.xml") or
            n.endswith("_VIC.xml")
            for n in names
        )

    if not has_cfx_xml:
        raise ValueError(
            "This ZIP file does not contain recognized CFX Opus XML exports.\n\n"
            "Expected files like:\n"
            "  - Allelic Discrimination Results_ADSheet.xml\n"
            "  - Quantification Amplification Results_FAM.xml\n\n"
            "Please ZIP the CFX Maestro XML export folder and upload."
        )

    return parse_cfx_xml_zip(file_path)


def _handle_quantstudio(file_path: str, filename: str) -> UnifiedData:
    """Handle QuantStudio 3 .xls files with smart detection."""
    from app.parsers.quantstudio import parse_quantstudio, parse_quantstudio_amplification

    wb = xlrd.open_workbook(file_path)
    sheet = wb.sheet_by_index(0)

    # Find header row
    header_row = None
    for r in range(min(60, sheet.nrows)):
        val = sheet.cell_value(r, 0)
        if isinstance(val, str) and val.strip().lower() == "well":
            header_row = r
            break

    if header_row is None:
        raise ValueError(
            "This doesn't appear to be a QuantStudio data file.\n"
            "Upload the 'Multicomponent Data' export file (.xls) from QuantStudio."
        )

    headers_upper = {
        sheet.cell_value(header_row, c).strip().upper()
        for c in range(sheet.ncols)
        if isinstance(sheet.cell_value(header_row, c), str)
    }

    # Check for Multicomponent Data (FAM + VIC/HEX + Cycle)
    has_fam = "FAM" in headers_upper
    has_allele2 = "VIC" in headers_upper or "HEX" in headers_upper
    has_cycle = "CYCLE" in headers_upper

    if has_fam and has_allele2 and has_cycle:
        return parse_quantstudio(file_path)

    # Check for Amplification Data (Target Name, Rn, Delta Rn)
    has_target = "TARGET NAME" in headers_upper
    has_rn = "RN" in headers_upper or "DELTA RN" in headers_upper

    if has_target and has_rn and has_cycle:
        return parse_quantstudio_amplification(file_path)

    # Identify what file this is and give helpful error
    file_type = _identify_qs_file(filename, headers_upper)
    _raise_qs_error(file_type, headers_upper)


def _identify_qs_file(filename: str, headers: set[str]) -> str:
    """Identify which QuantStudio export this is."""
    # Detect by headers first (more reliable than filename)
    # Order matters: Results has SNP ASSAY NAME too, so check Results before Sample Setup
    if "X1-M1" in headers or "X2-M2" in headers:
        return "Raw Data"
    if "CALL" in headers and ("ALLELE1 DELTA RN" in headers or "QUALITY(%)" in headers):
        return "Results"
    if "ALLELE1 REPORTER" in headers or ("SNP ASSAY NAME" in headers and "CALL" not in headers):
        return "Sample Setup"

    # Fallback to filename
    for key, name in QS_KNOWN_FILES.items():
        if key in filename:
            return name

    return "unknown"


def _raise_qs_error(file_type: str, headers: set[str]):
    """Raise a helpful error message for non-usable QuantStudio files."""
    hints = {
        "Raw Data": (
            "This is a 'Raw Data' file with pre-deconvolution detector signals (x-m channels).\n"
            "These values haven't been separated into individual dye channels yet."
        ),
        "Results": (
            "This is a 'Results' file with endpoint genotyping calls only.\n"
            "It doesn't contain the per-cycle fluorescence curves needed for scatter plots."
        ),
        "Sample Setup": (
            "This is a 'Sample Setup' file with well assignments and sample names.\n"
            "It doesn't contain fluorescence data."
        ),
    }

    hint = hints.get(file_type, f"This file has columns: {', '.join(sorted(headers))}")

    raise ValueError(
        f"{'This is a ' + repr(file_type) + ' file. ' if file_type != 'unknown' else ''}"
        f"{hint}\n\n"
        f"Please upload the 'Multicomponent Data' file instead.\n"
        f"In QuantStudio, export: File > Export > Multicomponent Data (.xls)"
    )


def _handle_cfx_opus(file_path: str, filename: str) -> UnifiedData:
    """Handle CFX Opus .xlsx files with smart detection."""
    from app.parsers.cfx_opus import parse_cfx_opus, parse_cfx_endpoint, parse_cfx_allelic
    from app.parsers.xlsx_fixer import needs_fixing, fix_cfx_xlsx

    # Fix broken xlsx if needed
    fixed_path = None
    work_path = file_path
    if needs_fixing(file_path):
        fixed_path = fix_cfx_xlsx(file_path)
        work_path = fixed_path

    try:
        import openpyxl
        wb = openpyxl.load_workbook(work_path, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception:
        if fixed_path and os.path.exists(fixed_path):
            os.remove(fixed_path)
        raise ValueError(
            "Could not read this .xlsx file.\n"
            "If this is from CFX Opus/Maestro, the file may be corrupted."
        )
    finally:
        if fixed_path and os.path.exists(fixed_path):
            os.remove(fixed_path)

    has_dye_sheets = "FAM" in sheet_names and ("HEX" in sheet_names or "VIC" in sheet_names)

    # 1. Check for Quantification Amplification Results (FAM/HEX/ROX wide format with Cycle column)
    if has_dye_sheets:
        # Peek at FAM sheet to determine format
        fixed_path2 = None
        if needs_fixing(file_path):
            fixed_path2 = fix_cfx_xlsx(file_path)
            peek_path = fixed_path2
        else:
            peek_path = file_path

        try:
            import openpyxl
            wb = openpyxl.load_workbook(peek_path, data_only=True)
            fam_ws = wb["FAM"]
            headers = [cell.value for cell in fam_ws[1]]
            wb.close()

            # Wide format: [None, 'Cycle', 'A1', 'A2', ...]
            if headers and len(headers) > 2 and headers[1] == "Cycle":
                return parse_cfx_opus(file_path)

            # Long format with End RFU: [None, 'Well', 'Fluor', 'Target', 'Content', 'Sample', 'End RFU', ...]
            headers_str = [str(h).upper() if h else "" for h in headers]
            if "WELL" in headers_str and "END RFU" in headers_str:
                return parse_cfx_endpoint(file_path)

            # Plate view format (Melt Curve, Quantification Plate View) - NOT usable
            # These have column headers as numbers 1-12 for plate columns
            if any(isinstance(h, (int, float)) and h in range(1, 13) for h in headers):
                file_type = _identify_cfx_file(filename, sheet_names)
                _raise_cfx_error(file_type or "Plate View", sheet_names, filename)

            # Unknown dye-sheet format
            file_type = _identify_cfx_file(filename, sheet_names)
            _raise_cfx_error(file_type, sheet_names, filename)

        finally:
            if fixed_path2 and os.path.exists(fixed_path2):
                os.remove(fixed_path2)

    # 2. Check for Allelic Discrimination Results (ADSheet)
    if "ADSheet" in sheet_names:
        return parse_cfx_allelic(file_path)

    # 3. Identify the file and give helpful error
    file_type = _identify_cfx_file(filename, sheet_names)
    _raise_cfx_error(file_type, sheet_names, filename)


def _identify_cfx_file(filename: str, sheet_names: list[str]) -> str:
    """Identify which CFX Opus export this is."""
    for key, name in CFX_KNOWN_FILES.items():
        if key in filename:
            return name

    if "ADSheet" in sheet_names:
        return "Allelic Discrimination Results"
    if "ANOVA" in sheet_names:
        return "ANOVA Results"
    if "Standard Curve Results" in sheet_names:
        return "Standard Curve Results"

    return "unknown"


def _raise_cfx_error(file_type: str, sheet_names: list[str], filename: str):
    """Raise a helpful error message for non-usable CFX Opus files."""
    hints = {
        "Quantification Cq Results": "This file contains Cq (cycle threshold) values only, not fluorescence curves.",
        "Quantification Summary": "This file contains a summary of Cq values, not fluorescence data.",
        "Melt Curve Plate View Results": "This file contains melt curve analysis data, not amplification fluorescence.",
        "Quantification Plate View Results": "This file contains a plate-view summary of Cq values.",
        "Gene Expression Results": "This file contains gene expression analysis results.",
        "ANOVA Results": "This file contains ANOVA statistical results.",
        "Standard Curve Results": "This file contains standard curve calibration data.",
        "Plate View": "This file contains a plate-view summary, not per-cycle fluorescence data.",
    }

    hint = hints.get(file_type, f"Sheets found: {', '.join(sheet_names)}")

    raise ValueError(
        f"{'This is a ' + repr(file_type) + ' file. ' if file_type != 'unknown' else ''}"
        f"{hint}\n\n"
        f"Please upload one of these CFX Opus export files:\n"
        f"  - 'Quantification Amplification Results' (best - has per-cycle data)\n"
        f"  - 'End Point Results' (endpoint fluorescence)\n"
        f"  - 'Allelic Discrimination Results' (allele calls + RFU)"
    )
