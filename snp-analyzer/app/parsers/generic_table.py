from __future__ import annotations

import csv
import math
import re
import zipfile
from pathlib import Path
from typing import Any

from app.assays.registry import validate_mapping_config
from app.config import WELL_COLS, WELL_ROWS
from app.import_errors import ImportErrorCode, ImportValidationError, make_issue, raise_import_error
from app.import_models import (
    AssayModeId,
    ImportPreview,
    ImportReading,
    ImportRole,
    ImportRun,
    MappingConfig,
    NormalizationMode,
    ReporterChannel,
    ValidationIssue,
)
from app.models import UnifiedData, WellCycleData
from app.parsers.detector import _validate_zip_archive


MAX_IMPORT_ROWS = 100_000
MAX_IMPORT_SHEETS = 20
MAX_IMPORT_WELLS = 384
MAX_IMPORT_CYCLES = 200
MAX_IMPORT_CHANNELS = 16

_WELL_RE = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")
_GENERIC_LONG_HEADERS = {"well", "cycle", "dye", "role", "rfu", "sample", "target", "sample_type"}
_GENERIC_WIDE_HEADERS = {"well", "cycle", "ch1_rfu", "ch2_rfu", "ch3_rfu", "ch4_rfu"}


class GenericTableParser:
    parser_id = "generic-table"

    def __init__(self, *, instrument: str = "Generic Mapped Table") -> None:
        self.instrument = instrument

    def sniff(self, file_path: Path, original_filename: str) -> bool:
        suffix = Path(original_filename or file_path.name).suffix.lower()
        if suffix not in {".csv", ".tsv", ".txt", ".xlsx"}:
            return False
        try:
            table = _read_table(file_path, _minimal_config())
        except ImportValidationError:
            return False
        headers = set(table.headers)
        return not (
            _GENERIC_LONG_HEADERS.issubset(headers)
            or _GENERIC_WIDE_HEADERS.issubset(headers)
        )

    def preview(self, file_path: Path, original_filename: str) -> ImportPreview:
        table = _read_table(file_path, MappingConfig(
            assay_mode=AssayModeId.WT_MT,
            channel_roles={"preview": ImportRole.UNKNOWN},
        ))
        headers = table.headers
        return ImportPreview(
            preview_id="",
            parser_id=self.parser_id,
            filename=original_filename,
            candidate_tables=[table.sheet_name],
            inferred_headers=headers,
            sample_rows=table.rows[:5],
            channel_candidates=[ReporterChannel(channel_id=header) for header in headers if "rfu" in header.lower()],
            warnings=[
                make_issue(
                    ImportErrorCode.MAPPING_CONFIG_REQUIRED,
                    message="Confirm table structure and channel-to-role mapping before import.",
                )
            ],
        )

    def parse(
        self,
        file_path: Path,
        original_filename: str,
        mapping_config: MappingConfig | None = None,
    ) -> ImportRun:
        if mapping_config is None:
            raise_import_error(ImportErrorCode.MAPPING_CONFIG_REQUIRED)

        table = _read_table(file_path, mapping_config)
        _validate_mapping_shape(table.headers, mapping_config)
        _raise_for_mapping_issues(validate_mapping_config(mapping_config).issues)

        if mapping_config.rfu_columns:
            return self._parse_channel_columns(table, mapping_config)
        return self._parse_dye_rows(table, mapping_config)

    def to_unified(self, import_run: ImportRun) -> UnifiedData:
        return _to_duplex_unified(import_run)

    def _parse_channel_columns(self, table: "_Table", config: MappingConfig) -> ImportRun:
        issues: list[ValidationIssue] = []
        channel_order = list(config.rfu_columns.keys())
        samples: dict[str, str] = {}
        targets: dict[str, str] = {}
        readings: list[ImportReading] = []
        seen: set[tuple[str, int, str]] = set()
        populated_channels: set[str] = set()

        for row_number, row in table.iter_data_rows():
            well_column = _column_name(config.well_column)
            cycle_column = _column_name(config.cycle_column)
            well = _parse_well(row.get(well_column, ""), row_number, config.well_column or "")
            cycle = _parse_cycle(row.get(cycle_column, ""), row_number, config.cycle_column or "")
            _capture_optional(samples, well, row, config.sample_column)
            _capture_optional(targets, well, row, config.target_column)

            for channel_id, column in config.rfu_columns.items():
                role = config.channel_roles.get(channel_id, ImportRole.UNKNOWN)
                if role in {ImportRole.EXCLUDED, ImportRole.UNKNOWN}:
                    continue
                normalized_column = _column_name(column)
                raw_value = row.get(normalized_column, "")
                if raw_value == "":
                    continue
                rfu = _parse_rfu(raw_value, row_number, column, config.decimal_separator or ".")
                key = (well, cycle, channel_id)
                if key in seen:
                    issues.append(_duplicate_issue(key, row_number, column))
                    continue
                seen.add(key)
                populated_channels.add(channel_id)
                readings.append(ImportReading(well=well, cycle=cycle, channel_id=channel_id, rfu=rfu))

        issues.extend(_missing_populated_channel_issues(config, populated_channels))
        _raise_if_issues(issues)
        return _build_import_run(self.instrument, config, channel_order, None, readings, samples, targets)

    def _parse_dye_rows(self, table: "_Table", config: MappingConfig) -> ImportRun:
        issues: list[ValidationIssue] = []
        samples: dict[str, str] = {}
        targets: dict[str, str] = {}
        readings: list[ImportReading] = []
        seen: set[tuple[str, int, str]] = set()
        populated_channels: set[str] = set()
        channel_order: list[str] = []
        dye_names: dict[str, str] = {}

        for row_number, row in table.iter_data_rows():
            well_column = _column_name(config.well_column)
            cycle_column = _column_name(config.cycle_column)
            dye_column = _column_name(config.dye_column)
            rfu_column = _column_name(config.rfu_column)
            well = _parse_well(row.get(well_column, ""), row_number, config.well_column or "")
            cycle = _parse_cycle(row.get(cycle_column, ""), row_number, config.cycle_column or "")
            channel_id = (row.get(dye_column, "") or "").strip()
            if not channel_id:
                issues.append(make_issue(ImportErrorCode.MISSING_FIELD, row=row_number, column=config.dye_column))
                continue
            if channel_id not in config.channel_roles:
                continue
            role = config.channel_roles[channel_id]
            if role in {ImportRole.EXCLUDED, ImportRole.UNKNOWN}:
                continue
            if channel_id not in channel_order:
                channel_order.append(channel_id)
            dye_names[channel_id] = channel_id
            _capture_optional(samples, well, row, config.sample_column)
            _capture_optional(targets, well, row, config.target_column)
            rfu = _parse_rfu(row.get(rfu_column, ""), row_number, config.rfu_column or "", config.decimal_separator or ".")
            key = (well, cycle, channel_id)
            if key in seen:
                issues.append(_duplicate_issue(key, row_number, config.rfu_column))
                continue
            seen.add(key)
            populated_channels.add(channel_id)
            readings.append(ImportReading(well=well, cycle=cycle, channel_id=channel_id, rfu=rfu))

        issues.extend(_missing_populated_channel_issues(config, populated_channels))
        _raise_if_issues(issues)
        return _build_import_run(self.instrument, config, channel_order, dye_names, readings, samples, targets)


class GenericLongParser(GenericTableParser):
    parser_id = "generic-long"

    def __init__(self) -> None:
        super().__init__(instrument="Generic Long Table")

    def sniff(self, file_path: Path, original_filename: str) -> bool:
        try:
            table = _read_table(file_path, _minimal_config())
        except ImportValidationError:
            return False
        return _GENERIC_LONG_HEADERS.issubset(set(table.headers))

    def parse(
        self,
        file_path: Path,
        original_filename: str,
        mapping_config: MappingConfig | None = None,
    ) -> ImportRun:
        table = _read_table(file_path, _minimal_config())
        headers = set(table.headers)
        if "cq" in headers and "rfu" not in headers:
            raise_import_error(ImportErrorCode.CQ_ENDPOINT_ONLY)
        _validate_required_headers(table.headers, _GENERIC_LONG_HEADERS)

        channel_roles: dict[str, ImportRole] = {}
        for _, row in table.iter_data_rows():
            dye = row.get("dye", "").strip()
            raw_role = row.get("role", "").strip()
            if dye and raw_role:
                channel_roles[dye] = ImportRole(raw_role)

        config = MappingConfig(
            assay_mode=_infer_assay_mode(set(channel_roles.values())),
            normalization_mode=(
                NormalizationMode.PASSIVE_REFERENCE
                if ImportRole.NORMALIZATION in set(channel_roles.values())
                else NormalizationMode.NONE
            ),
            channel_roles=channel_roles or {"unknown": ImportRole.UNKNOWN},
            delimiter=table.delimiter,
            decimal_separator=table.decimal_separator,
            header_row=0,
            first_data_row=1,
            well_column="well",
            cycle_column="cycle",
            sample_column="sample",
            target_column="target",
            dye_column="dye",
            role_column="role",
            rfu_column="rfu",
        )
        _raise_for_mapping_issues(validate_mapping_config(config).issues)
        return self._parse_dye_rows(table, config)


class GenericWideParser(GenericTableParser):
    parser_id = "generic-wide"

    def __init__(self) -> None:
        super().__init__(instrument="Generic Wide Table")

    def sniff(self, file_path: Path, original_filename: str) -> bool:
        try:
            table = _read_table(file_path, _minimal_config())
        except ImportValidationError:
            return False
        return _GENERIC_WIDE_HEADERS.issubset(set(table.headers))

    def parse(
        self,
        file_path: Path,
        original_filename: str,
        mapping_config: MappingConfig | None = None,
    ) -> ImportRun:
        if mapping_config is None:
            raise_import_error(ImportErrorCode.MAPPING_CONFIG_REQUIRED)
        table = _read_table(file_path, mapping_config)
        _validate_required_headers(table.headers, _GENERIC_WIDE_HEADERS)
        return super().parse(file_path, original_filename, mapping_config)


class _Table:
    def __init__(
        self,
        *,
        headers: list[str],
        rows: list[dict[str, str]],
        delimiter: str,
        decimal_separator: str,
        sheet_name: str = "Sheet1",
        first_data_row: int = 1,
    ) -> None:
        self.headers = headers
        self.rows = rows
        self.delimiter = delimiter
        self.decimal_separator = decimal_separator
        self.sheet_name = sheet_name
        self.first_data_row = first_data_row

    def iter_data_rows(self) -> list[tuple[int, dict[str, str]]]:
        return [(self.first_data_row + index + 1, row) for index, row in enumerate(self.rows)]


def _minimal_config() -> MappingConfig:
    return MappingConfig(assay_mode=AssayModeId.WT_MT, channel_roles={"preview": ImportRole.UNKNOWN})


def _read_table(file_path: Path, config: MappingConfig) -> _Table:
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx_table(file_path, config)
    return _read_delimited_table(file_path, config)


def _read_delimited_table(file_path: Path, config: MappingConfig) -> _Table:
    text = file_path.read_text(encoding="utf-8-sig")
    sample = text[:4096]
    delimiter = config.delimiter or _detect_delimiter(sample, file_path.suffix)
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    if not rows:
        raise_import_error(ImportErrorCode.UNSUPPORTED_CONTENT)
    return _table_from_matrix(rows, config, delimiter=delimiter, sheet_name=file_path.name)


def _read_xlsx_table(file_path: Path, config: MappingConfig) -> _Table:
    if not zipfile.is_zipfile(file_path):
        raise_import_error(ImportErrorCode.UNSUPPORTED_CONTENT, message="XLSX file is not a valid ZIP archive.")
    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            _validate_zip_archive(zf)
    except ValueError as exc:
        raise_import_error(ImportErrorCode.FILE_LIMIT_EXCEEDED, message=str(exc))
    import openpyxl

    workbook = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    try:
        if len(workbook.sheetnames) > MAX_IMPORT_SHEETS:
            raise_import_error(ImportErrorCode.FILE_LIMIT_EXCEEDED)
        worksheet = workbook[workbook.sheetnames[0]]
        matrix = [
            ["" if value is None else str(value) for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()
    return _table_from_matrix(matrix, config, delimiter=",", sheet_name=workbook.sheetnames[0])


def _table_from_matrix(
    matrix: list[list[str]],
    config: MappingConfig,
    *,
    delimiter: str,
    sheet_name: str,
) -> _Table:
    header_index = config.header_row or 0
    first_data_index = config.first_data_row if config.first_data_row is not None else header_index + 1
    if header_index >= len(matrix):
        raise_import_error(ImportErrorCode.MISSING_FIELD, message="Header row is outside the table.")
    headers = [_normalize_header(value) for value in matrix[header_index]]
    if len(matrix) - first_data_index > MAX_IMPORT_ROWS:
        raise_import_error(ImportErrorCode.FILE_LIMIT_EXCEEDED)
    if "cq" in headers and "rfu" not in headers and not any(header.endswith("_rfu") for header in headers):
        raise_import_error(ImportErrorCode.CQ_ENDPOINT_ONLY)

    rows: list[dict[str, str]] = []
    for row in matrix[first_data_index:]:
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        rows.append({header: str(value).strip() for header, value in zip(headers, padded) if header})
    decimal_separator = config.decimal_separator or _infer_decimal_separator(rows)
    return _Table(
        headers=headers,
        rows=rows,
        delimiter=delimiter,
        decimal_separator=decimal_separator,
        sheet_name=sheet_name,
        first_data_row=first_data_index,
    )


def _detect_delimiter(sample: str, suffix: str) -> str:
    if suffix.lower() in {".tsv", ".txt"} and "\t" in sample:
        return "\t"
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        return "\t" if suffix.lower() == ".tsv" else ","


def _infer_decimal_separator(rows: list[dict[str, str]]) -> str:
    for row in rows[:25]:
        for value in row.values():
            if re.fullmatch(r"-?\d+,\d+", value):
                return ","
    return "."


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _column_name(value: str | None) -> str:
    return _normalize_header(value)


def _validate_mapping_shape(headers: list[str], config: MappingConfig) -> None:
    issues: list[ValidationIssue] = []
    for column in [config.well_column, config.cycle_column]:
        if not column or _column_name(column) not in headers:
            issues.append(make_issue(ImportErrorCode.MISSING_FIELD, column=column))
    if config.rfu_columns:
        for column in config.rfu_columns.values():
            if _column_name(column) not in headers:
                issues.append(make_issue(ImportErrorCode.MISSING_FIELD, column=column))
    elif not config.dye_column or not config.rfu_column:
        issues.append(make_issue(ImportErrorCode.MISSING_FIELD, message="RFU and dye columns are required."))
    else:
        for column in [config.dye_column, config.rfu_column]:
            if _column_name(column) not in headers:
                issues.append(make_issue(ImportErrorCode.MISSING_FIELD, column=column))
    _raise_if_issues(issues)


def _validate_required_headers(headers: list[str], required: set[str]) -> None:
    missing = sorted(required - set(headers))
    if missing:
        raise_import_error(
            ImportErrorCode.MISSING_FIELD,
            message=f"Missing required headers: {', '.join(missing)}",
            context={"headers": missing},
        )


def _parse_well(value: str, row_number: int, column: str) -> str:
    well = value.strip().upper()
    match = _WELL_RE.match(well)
    if not match:
        raise_import_error(ImportErrorCode.MALFORMED_WELL, row=row_number, column=column, context={"well": value})
    row_label, col_text = match.groups()
    if row_label not in WELL_ROWS or int(col_text) < 1 or int(col_text) > WELL_COLS:
        raise_import_error(ImportErrorCode.MALFORMED_WELL, row=row_number, column=column, context={"well": value})
    return well


def _parse_cycle(value: str, row_number: int, column: str) -> int:
    try:
        cycle = int(str(value).strip())
    except ValueError:
        raise_import_error(ImportErrorCode.INVALID_NUMERIC_VALUE, row=row_number, column=column)
    if cycle < 1:
        raise_import_error(ImportErrorCode.INVALID_NUMERIC_VALUE, row=row_number, column=column)
    return cycle


def _parse_rfu(value: str, row_number: int, column: str, decimal_separator: str) -> float:
    text = str(value).strip()
    if not text:
        raise_import_error(ImportErrorCode.INVALID_NUMERIC_VALUE, row=row_number, column=column)
    if text.startswith("="):
        raise_import_error(ImportErrorCode.FORMULA_AS_RFU, row=row_number, column=column)
    if decimal_separator == "." and re.fullmatch(r"-?\d+,\d+", text):
        raise_import_error(ImportErrorCode.DECIMAL_SEPARATOR_MISMATCH, row=row_number, column=column)
    normalized = text.replace(",", ".") if decimal_separator == "," else text
    try:
        rfu = float(normalized)
    except ValueError:
        raise_import_error(ImportErrorCode.INVALID_NUMERIC_VALUE, row=row_number, column=column)
    if not math.isfinite(rfu):
        raise_import_error(ImportErrorCode.INVALID_NUMERIC_VALUE, row=row_number, column=column)
    return rfu


def _capture_optional(values: dict[str, str], well: str, row: dict[str, str], column: str | None) -> None:
    normalized_column = _column_name(column)
    if column and row.get(normalized_column):
        values.setdefault(well, row[normalized_column])


def _missing_populated_channel_issues(config: MappingConfig, populated_channels: set[str]) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for channel_id, role in config.channel_roles.items():
        if role == ImportRole.NORMALIZATION and channel_id not in populated_channels:
            issues.append(make_issue(ImportErrorCode.MISSING_NORMALIZATION_CHANNEL, channel_id=channel_id))
        elif role in {ImportRole.WT, ImportRole.MT1, ImportRole.MT2, ImportRole.MT3} and channel_id not in populated_channels:
            issues.append(make_issue(ImportErrorCode.MISSING_REQUIRED_ROLE, channel_id=channel_id))
    return issues


def _build_import_run(
    instrument: str,
    config: MappingConfig,
    channel_order: list[str],
    dye_names: dict[str, str] | None,
    readings: list[ImportReading],
    samples: dict[str, str],
    targets: dict[str, str],
) -> ImportRun:
    _validate_import_caps(readings, channel_order)
    channels = [
        ReporterChannel(
            channel_id=channel_id,
            dye_name=(dye_names or {}).get(channel_id),
            role=config.channel_roles.get(channel_id, ImportRole.UNKNOWN),
        )
        for channel_id in channel_order
        if config.channel_roles.get(channel_id) not in {ImportRole.EXCLUDED, ImportRole.UNKNOWN}
    ]
    return ImportRun(
        instrument=instrument,
        reporter_channels=channels,
        readings=readings,
        samples=samples,
        targets=targets,
        metadata={
            "assay_mode": config.assay_mode.value,
            "normalization_mode": config.normalization_mode.value,
        },
    )


def _validate_import_caps(readings: list[ImportReading], channels: list[str]) -> None:
    if len({reading.well for reading in readings}) > MAX_IMPORT_WELLS:
        raise_import_error(ImportErrorCode.FILE_LIMIT_EXCEEDED)
    if len({reading.cycle for reading in readings}) > MAX_IMPORT_CYCLES:
        raise_import_error(ImportErrorCode.FILE_LIMIT_EXCEEDED)
    if len(channels) > MAX_IMPORT_CHANNELS:
        raise_import_error(ImportErrorCode.FILE_LIMIT_EXCEEDED)


def _infer_assay_mode(roles: set[ImportRole]) -> AssayModeId:
    if ImportRole.MT3 in roles:
        return AssayModeId.WT_MT1_MT2_MT3
    if ImportRole.MT2 in roles:
        return AssayModeId.WT_MT1_MT2
    return AssayModeId.WT_MT


def _raise_for_mapping_issues(issues: list[ValidationIssue]) -> None:
    if not issues:
        return
    adjusted: list[ValidationIssue] = []
    for issue in issues:
        policy_code = ImportErrorCode(issue.code)
        adjusted.append(
            make_issue(
                policy_code,
                message=issue.message,
                row=issue.row,
                column=issue.column,
                channel_id=issue.channel_id,
                context=issue.context,
            )
        )
    raise ImportValidationError(adjusted)


def _raise_if_issues(issues: list[ValidationIssue]) -> None:
    if issues:
        raise ImportValidationError(issues)


def _duplicate_issue(key: tuple[str, int, str], row_number: int, column: str | None) -> ValidationIssue:
    well, cycle, channel_id = key
    return make_issue(
        ImportErrorCode.DUPLICATE_READING,
        row=row_number,
        column=column,
        channel_id=channel_id,
        context={"well": well, "cycle": cycle, "channel_id": channel_id},
    )


def _to_duplex_unified(import_run: ImportRun) -> UnifiedData:
    channels_by_role = {channel.role: channel for channel in import_run.reporter_channels}
    wt = channels_by_role.get(ImportRole.WT)
    mt1 = channels_by_role.get(ImportRole.MT1)
    if wt is None or mt1 is None:
        raise ValueError("WT and MT1 channels are required for legacy UnifiedData conversion")
    normalization = channels_by_role.get(ImportRole.NORMALIZATION)
    grouped: dict[tuple[str, int], dict[str, float]] = {}
    for reading in import_run.readings:
        grouped.setdefault((reading.well, reading.cycle), {})[reading.channel_id] = reading.rfu
    data: list[WellCycleData] = []
    for (well, cycle), values in sorted(grouped.items()):
        if wt.channel_id not in values or mt1.channel_id not in values:
            continue
        data.append(
            WellCycleData(
                well=well,
                cycle=cycle,
                fam=values[wt.channel_id],
                allele2=values[mt1.channel_id],
                rox=values.get(normalization.channel_id) if normalization else None,
            )
        )
    return UnifiedData(
        instrument=import_run.instrument,
        allele2_dye=mt1.dye_name or mt1.channel_id,
        wells=sorted({reading.well for reading in import_run.readings}),
        cycles=sorted({reading.cycle for reading in import_run.readings}),
        data=data,
        has_rox=normalization is not None,
        sample_names=import_run.samples or None,
    )
