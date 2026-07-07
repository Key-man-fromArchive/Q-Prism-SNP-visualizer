"""Build an XLSX workbook of the analysis: a Summary sheet with the embedded
allele-discrimination plot + metadata, and a Results sheet with the full
genotype table (editable data)."""
from __future__ import annotations

import io
from datetime import datetime, timezone


def build_xlsx(
    *,
    instrument: str,
    allele2_dye: str,
    cycle: int,
    filename: str,
    scatter_points: list[dict],
    table_headers: list[str],
    table_rows: list[list],
    genotype_counts: dict[str, int],
    qc: dict[str, object] | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font
    from app.reporting.charts import render_scatter_png

    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "SNP Allele Discrimination Report"
    ws["A1"].font = Font(size=14, bold=True)

    meta = [
        ("File", filename or "-"),
        ("Instrument", instrument),
        ("Allele 2 dye", allele2_dye),
        ("Analysis cycle", cycle),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    row = 3
    for key, value in meta:
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Genotype counts").font = Font(bold=True)
    row += 1
    for gt, count in genotype_counts.items():
        ws.cell(row=row, column=1, value=gt)
        ws.cell(row=row, column=2, value=count)
        row += 1

    if qc:
        row += 1
        ws.cell(row=row, column=1, value="QC").font = Font(bold=True)
        row += 1
        for key, value in qc.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24

    # Embed the allele-discrimination scatter (matplotlib PNG) next to the meta.
    try:
        png = render_scatter_png(scatter_points, allele2_dye)
        img = XLImage(io.BytesIO(png))
        ws.add_image(img, "D3")
    except Exception:
        # If plotting fails, still deliver the data-only workbook.
        pass

    # --- Results sheet ---
    ws2 = wb.create_sheet("Results")
    ws2.append(table_headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in table_rows:
        ws2.append(r)
    ws2.freeze_panes = "A2"
    for col in ("A", "B", "C", "D"):
        ws2.column_dimensions[col].width = 16

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
