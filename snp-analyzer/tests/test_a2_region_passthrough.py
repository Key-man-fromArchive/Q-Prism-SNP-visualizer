"""A2: silent-regression fix -- downstream consumers (statistics, export, qc,
asg_result) must become multi-marker-aware.

When ``cluster_store[sid].regions`` is set, the plate holds MULTIPLE markers,
each with its own ploidy/assignments/genotype_counts. The flat top-level
``assignments`` is only a merge and the top-level ``ploidy`` is legacy/session
-level, NOT authoritative for a multi-marker session. These tests pin:

1. Per-marker breakdown when ``regions`` is set (statistics / export / qc).
2. BYTE-IDENTICAL single-marker (``regions is None``) behavior -- the historical
   golden path must be completely unaffected.
3. ``asg_result.build_result_snapshot`` refuses (409) to emit a half-formed
   plate-level snapshot for a multi-marker session; single-marker snapshots
   are unchanged.
"""
import csv
import io
import os
from types import SimpleNamespace
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient

from app.auth import TokenData, get_current_user
from app.models import (
    ClusteringResult,
    RegionResult,
    UnifiedData,
    WellCycleData,
)


def _unified_single_marker() -> UnifiedData:
    """6 wells, one marker, diploid -- the historical golden path."""
    wells = [f"A{i}" for i in range(1, 7)]
    data = []
    for i, w in enumerate(wells):
        fam, a2 = (0.9, 0.1) if i < 3 else (0.1, 0.9)
        data.append(WellCycleData(well=w, cycle=1, fam=fam, allele2=a2, rox=None))
    return UnifiedData(
        instrument="QuantStudio 3", allele2_dye="VIC", wells=wells,
        cycles=[1], data=data, has_rox=False,
    )


def _unified_multi_marker() -> UnifiedData:
    """8 wells split across two markers: A1-A4 (diploid), B1-B4 (tetraploid)."""
    wells = [f"A{i}" for i in range(1, 5)] + [f"B{i}" for i in range(1, 5)]
    data = []
    # Marker "diploidM": A1,A2 Allele1Homo-ish; A3,A4 Allele2Homo-ish.
    fam_a2 = [(0.9, 0.1), (0.9, 0.1), (0.1, 0.9), (0.1, 0.9)]
    for w, (fam, a2) in zip(wells[:4], fam_a2):
        data.append(WellCycleData(well=w, cycle=1, fam=fam, allele2=a2, rox=None))
    # Marker "tetraM": B1..B4 tetraploid dosage classes.
    fam_a2_b = [(0.95, 0.05), (0.7, 0.3), (0.3, 0.7), (0.05, 0.95)]
    for w, (fam, a2) in zip(wells[4:], fam_a2_b):
        data.append(WellCycleData(well=w, cycle=1, fam=fam, allele2=a2, rox=None))
    return UnifiedData(
        instrument="QuantStudio 3", allele2_dye="VIC", wells=wells,
        cycles=[1], data=data, has_rox=False,
    )


def _single_marker_cluster_result() -> ClusteringResult:
    return ClusteringResult(
        algorithm="threshold",
        cycle=1,
        assignments={
            "A1": "Allele 1 Homo", "A2": "Allele 1 Homo", "A3": "Allele 1 Homo",
            "A4": "Allele 2 Homo", "A5": "Allele 2 Homo", "A6": "Allele 2 Homo",
        },
        ploidy=2,
    )


def _multi_marker_cluster_result() -> ClusteringResult:
    region_a = RegionResult(
        id="m1", name="diploidM", wells=["A1", "A2", "A3", "A4"], ploidy=2,
        assignments={
            "A1": "Allele 1 Homo", "A2": "Allele 1 Homo",
            "A3": "Allele 2 Homo", "A4": "Allele 2 Homo",
        },
        genotype_counts={"AA": 2, "AB": 0, "BB": 2, "excluded": 0},
    )
    region_b = RegionResult(
        id="m2", name="tetraM", wells=["B1", "B2", "B3", "B4"], ploidy=4,
        assignments={
            "B1": "AAAA", "B2": "AAAB", "B3": "ABBB", "B4": "BBBB",
        },
        genotype_counts={"AAAA": 1, "AAAB": 1, "AABB": 0, "ABBB": 1, "BBBB": 1, "excluded": 0},
    )
    flat_assignments = {**region_a.assignments, **region_b.assignments}
    return ClusteringResult(
        algorithm="threshold",
        cycle=1,
        assignments=flat_assignments,
        ploidy=2,  # legacy session ploidy -- NOT authoritative when regions is set
        regions=[region_a, region_b],
    )


@pytest.fixture
def client(tmp_path):
    env = patch.dict(os.environ, {
        "JWT_SECRET_KEY": "test-secret-that-is-long-enough-for-a2-passthrough",
        "ADMIN_PASSWORD": "StrongerOperatorPassword123!",
        "SNP_AUTH_MODE": "local",
    }, clear=False)
    env.start()
    import app.db as db
    if db._conn is not None:
        db._conn.close()
    db._conn = None
    db.DB_PATH = tmp_path / "t.sqlite3"
    from app.main import app
    from app.routers import upload, clustering

    async def override():
        return TokenData(user_id="u", username="u", role="user")
    app.dependency_overrides[get_current_user] = override
    upload.sessions.clear(); clustering.welltype_store.clear(); clustering.cluster_store.clear()
    with TestClient(app) as c:
        yield SimpleNamespace(client=c, upload=upload, clustering=clustering, db=db)
    app.dependency_overrides.pop(get_current_user, None)
    upload.sessions.clear(); clustering.welltype_store.clear(); clustering.cluster_store.clear()
    if db._conn is not None:
        db._conn.close()
    db._conn = None
    env.stop()


def _register(client, sid, unified):
    client.upload.sessions[sid] = unified
    client.db.save_session(sid, unified, filename="t.eds", user_id=None)


# ---------------------------------------------------------------------------
# 1. statistics.py
# ---------------------------------------------------------------------------

def test_statistics_single_marker_golden_path_unchanged(client):
    _register(client, "s1", _unified_single_marker())
    client.clustering.cluster_store["s1"] = _single_marker_cluster_result()

    resp = client.client.get("/api/data/s1/statistics")
    assert resp.status_code == 200, resp.text
    body = resp.json()

    assert "markers" not in body
    assert body["allele_frequency"]["p"] == pytest.approx(0.5)
    assert body["genotype_distribution"] == {"Allele 1 Homo": 3, "Allele 2 Homo": 3}
    assert body["total_wells"] == 6


def test_statistics_multi_marker_reports_per_marker_breakdown(client):
    _register(client, "s2", _unified_multi_marker())
    client.clustering.cluster_store["s2"] = _multi_marker_cluster_result()

    resp = client.client.get("/api/data/s2/statistics")
    assert resp.status_code == 200, resp.text
    body = resp.json()

    assert "markers" in body
    assert len(body["markers"]) == 2
    by_id = {m["id"]: m for m in body["markers"]}

    diploid_m = by_id["m1"]
    assert diploid_m["ploidy"] == 2
    assert diploid_m["genotype_counts"]["AA"] == 2
    assert diploid_m["genotype_counts"]["BB"] == 2
    assert diploid_m["total_wells"] == 4
    # Real diploid HWE/allele-frequency computed from THIS marker's own counts,
    # not the flat plate-level pool.
    assert diploid_m["allele_frequency"]["p"] == pytest.approx(0.5)

    tetra_m = by_id["m2"]
    assert tetra_m["ploidy"] == 4
    assert tetra_m["genotype_counts"]["AAAA"] == 1
    assert tetra_m["genotype_counts"]["BBBB"] == 1
    assert tetra_m["total_wells"] == 4
    # Dosage-keyed counts, not collapsed into diploid AA/AB/BB.
    assert "AA" not in tetra_m["genotype_counts"]


def test_statistics_multi_marker_manual_override_applies_per_marker(client):
    _register(client, "s3", _unified_multi_marker())
    client.clustering.cluster_store["s3"] = _multi_marker_cluster_result()
    client.clustering.welltype_store["s3"] = {"A1": "Allele 2 Homo"}

    resp = client.client.get("/api/data/s3/statistics")
    body = resp.json()
    by_id = {m["id"]: m for m in body["markers"]}
    assert by_id["m1"]["genotype_counts"]["AA"] == 1  # A1 manually overridden away
    assert by_id["m1"]["genotype_counts"]["BB"] == 3


# ---------------------------------------------------------------------------
# 2. export.py (CSV)
# ---------------------------------------------------------------------------

def test_export_csv_single_marker_unchanged(client):
    _register(client, "s1", _unified_single_marker())
    client.clustering.cluster_store["s1"] = _single_marker_cluster_result()

    resp = client.client.get("/api/data/s1/export/csv?cycle=1&use_rox=false")
    assert resp.status_code == 200, resp.text
    reader = csv.reader(io.StringIO(resp.text))
    rows = list(reader)
    header = rows[0]
    assert "Marker" not in header
    assert header == [
        "Well", "Sample Name", "Genotype", "Confidence (%)",
        "FAM (norm)", "VIC (norm)", "FAM (raw)", "VIC (raw)", "ROX (raw)",
    ]
    assert len(rows) == 7  # header + 6 wells


def test_export_csv_multi_marker_has_marker_column_and_per_marker_vocab(client):
    _register(client, "s2", _unified_multi_marker())
    client.clustering.cluster_store["s2"] = _multi_marker_cluster_result()

    resp = client.client.get("/api/data/s2/export/csv?cycle=1&use_rox=false")
    assert resp.status_code == 200, resp.text
    reader = csv.reader(io.StringIO(resp.text))
    rows = list(reader)
    header = rows[0]
    assert "Marker" in header
    marker_idx = header.index("Marker")
    genotype_idx = header.index("Genotype")

    by_well = {r[header.index("Well")]: r for r in rows[1:]}
    assert by_well["A1"][marker_idx] == "diploidM"
    assert by_well["A1"][genotype_idx] == "Allele 1 Homo"
    assert by_well["B1"][marker_idx] == "tetraM"
    assert by_well["B1"][genotype_idx] == "AAAA"  # tetraploid vocab, not AA/AB/BB


# ---------------------------------------------------------------------------
# 3. export.py (XLSX)
# ---------------------------------------------------------------------------

def test_export_xlsx_single_marker_unchanged(client):
    _register(client, "s1", _unified_single_marker())
    client.clustering.cluster_store["s1"] = _single_marker_cluster_result()

    resp = client.client.get("/api/data/s1/export/xlsx?use_rox=false")
    assert resp.status_code == 200, resp.text

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    header = [c.value for c in wb["Results"][1]]
    assert "Marker" not in header
    assert wb["Results"].max_row == 7  # header + 6 wells


def test_export_xlsx_multi_marker_has_marker_column_and_per_marker_counts(client):
    _register(client, "s2", _unified_multi_marker())
    client.clustering.cluster_store["s2"] = _multi_marker_cluster_result()

    resp = client.client.get("/api/data/s2/export/xlsx?use_rox=false")
    assert resp.status_code == 200, resp.text

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    header = [c.value for c in wb["Results"][1]]
    assert "Marker" in header

    # Summary sheet genotype-counts section should carry a per-marker breakdown
    # (marker-prefixed keys), not one pooled diploid/tetraploid-mixed count.
    summary_values = [c.value for row in wb["Summary"].iter_rows() for c in row if c.value is not None]
    assert any("diploidM" in str(v) for v in summary_values)
    assert any("tetraM" in str(v) for v in summary_values)


# ---------------------------------------------------------------------------
# 4. qc.py
# ---------------------------------------------------------------------------

def test_qc_single_marker_golden_path_unchanged(client):
    _register(client, "s1", _unified_single_marker())
    client.clustering.cluster_store["s1"] = _single_marker_cluster_result()

    resp = client.client.get("/api/data/s1/qc?cycle=1&use_rox=false")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert "markers" not in body
    assert body["n_total"] == 6
    assert body["n_called"] == 6


def test_qc_multi_marker_reports_per_marker_cluster_separation(client):
    _register(client, "s2", _unified_multi_marker())
    client.clustering.cluster_store["s2"] = _multi_marker_cluster_result()

    resp = client.client.get("/api/data/s2/qc?cycle=1&use_rox=false")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert "markers" in body
    assert len(body["markers"]) == 2
    by_id = {m["id"]: m for m in body["markers"]}
    assert by_id["m1"]["ploidy"] == 2
    assert by_id["m1"]["n_total"] == 4
    assert by_id["m2"]["ploidy"] == 4
    assert by_id["m2"]["n_total"] == 4
    # Both markers independently resolve >= 2 clusters -> separation computable.
    assert by_id["m1"]["cluster_separation"] is not None
    assert by_id["m2"]["cluster_separation"] is not None


# ---------------------------------------------------------------------------
# 5. asg_result.py::build_result_snapshot
# ---------------------------------------------------------------------------

@pytest.fixture
def asg_env(tmp_path):
    env = patch.dict(os.environ, {
        "SNP_AUTH_MODE": "asg_launch",
        "JWT_SECRET_KEY": "test-secret-that-is-long-enough-for-a2-asg",
        "ASG_SNP_SERVICE_SECRET": "secret",
        "ASG_BASE_URL": "http://asg.local",
    }, clear=False)
    env.start()
    import app.db as db
    if db._conn is not None:
        db._conn.close()
    db._conn = None
    db.DB_PATH = tmp_path / "asg.sqlite3"
    yield db
    if db._conn is not None:
        db._conn.close()
    db._conn = None
    for module_name in ("app.routers.upload", "app.routers.clustering", "app.asg_session"):
        module = __import__(module_name, fromlist=["dummy"])
        getattr(module, "sessions", {}).clear()
        getattr(module, "cluster_store", {}).clear()
        getattr(module, "welltype_store", {}).clear()
        if hasattr(module, "clear_asg_launch_state"):
            module.clear_asg_launch_state()
    env.stop()


def _bind_asg_session(db, sid, unified):
    from app.asg_client import ASGLaunchContext, ASGLaunchSaveCredential
    from app.asg_session import bind_session_to_current_asg_launch, remember_asg_launch
    from app.db import init_db, save_session
    from app.routers.upload import sessions

    init_db()
    db.get_db().execute(
        "INSERT INTO users (id, username, hashed_password, display_name, role) VALUES (?, ?, ?, ?, ?)",
        ("asg-1", "owner@example.com", "!test", "Owner", "user"),
    )
    db.get_db().commit()
    sessions[sid] = unified
    save_session(sid, unified, filename="plate.xlsx", user_id="asg-1")
    remember_asg_launch(
        "asg-1",
        ASGLaunchContext("design_run_item", "101", {"marker_id": "M101"}),
        ASGLaunchSaveCredential("launch-1", "save-secret"),
        ["snp:save_result"],
        None,
    )
    bind_session_to_current_asg_launch(sid, "asg-1")


def test_asg_snapshot_single_marker_unchanged(asg_env):
    from app.asg_result import build_result_snapshot
    from app.routers.clustering import cluster_store

    _bind_asg_session(asg_env, "sid-single", _unified_single_marker())
    cluster_store["sid-single"] = _single_marker_cluster_result()

    snapshot = build_result_snapshot("sid-single", selected_cycle=1)
    assert snapshot["schema_version"] == 1
    assert snapshot["summary"]["genotype_counts"]["AA"] == 3
    assert snapshot["summary"]["genotype_counts"]["BB"] == 3
    assert len(snapshot["result"]["wells"]) == 6


def test_asg_snapshot_multi_marker_refuses_with_409(asg_env):
    from fastapi import HTTPException

    from app.asg_result import build_result_snapshot
    from app.routers.clustering import cluster_store

    _bind_asg_session(asg_env, "sid-multi", _unified_multi_marker())
    cluster_store["sid-multi"] = _multi_marker_cluster_result()

    with pytest.raises(HTTPException) as exc:
        build_result_snapshot("sid-multi", selected_cycle=1)
    assert exc.value.status_code == 409
    assert "schema_version 3" in exc.value.detail
