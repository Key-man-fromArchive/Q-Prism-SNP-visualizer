"""The XLSX export returns a valid workbook (Summary + Results) with the plot."""

import io
import os
from types import SimpleNamespace
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient

from app.auth import TokenData, get_current_user
from app.models import UnifiedData, WellCycleData


def _unified() -> UnifiedData:
    wells = [f"A{i}" for i in range(1, 7)]
    data = []
    for i, w in enumerate(wells):
        fam, a2 = (0.9, 0.1) if i < 3 else (0.1, 0.9)
        data.append(WellCycleData(well=w, cycle=1, fam=fam, allele2=a2, rox=None))
    return UnifiedData(
        instrument="QuantStudio 3", allele2_dye="VIC", wells=wells,
        cycles=[1], data=data, has_rox=False,
    )


@pytest.fixture
def client(tmp_path):
    env = patch.dict(os.environ, {
        "JWT_SECRET_KEY": "test-secret-that-is-long-enough-for-xlsx-export",
        "ADMIN_PASSWORD": "StrongerOperatorPassword123!",
        "SNP_AUTH_MODE": "local",
    }, clear=False)
    env.start()
    import app.db as db
    if db._conn is not None:
        db._conn.close()
    db._conn = None
    db.DB_PATH = tmp_path / "t.sqlite3"
    from app.main import app
    from app.routers import upload

    async def override():
        return TokenData(user_id="u", username="u", role="user")
    app.dependency_overrides[get_current_user] = override
    upload.sessions.clear()
    with TestClient(app) as c:
        yield SimpleNamespace(client=c, upload=upload, db=db)
    app.dependency_overrides.pop(get_current_user, None)
    upload.sessions.clear()
    if db._conn is not None:
        db._conn.close()
    db._conn = None
    env.stop()


def test_export_xlsx_returns_valid_workbook(client):
    unified = _unified()
    client.upload.sessions["s1"] = unified
    client.db.save_session("s1", unified, filename="t.eds", user_id=None)

    resp = client.client.get("/api/data/s1/export/xlsx?use_rox=false")
    assert resp.status_code == 200, resp.text
    assert resp.content[:2] == b"PK"  # xlsx is a zip
    assert "spreadsheetml" in resp.headers["content-type"]

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Summary", "Results"]
    assert wb["Results"].max_row == 7  # header + 6 wells
